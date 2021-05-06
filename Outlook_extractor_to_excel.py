import win32com.client
import os
from datetime import datetime, timedelta
import openpyxl as xl
import pandas as pd 
from openpyxl import load_workbook
def hello_world():
    print("welcome to Henry Outlook email reader")
    print("please be patient....")

def outlook_email():
    outlook=win32com.client.Dispatch('outlook.application') # get connected to the windows app name
    mapi= outlook.GetNamespace("MAPI")
    for account in mapi.Accounts:
        print(account.DeliveryStore.DisplayName)
    inbox=mapi.GetDefaultFolder(6) # the number is equal to the file name type , check windows to understand 
    # inbox=mapi.GetDefaultFolder(6).Folders["your_sub_folder"]
    message= inbox.Items
    #validated_mails=mapi.GetDefaultFolder(6).Folders("Done")
    
    #dataframes
    df = pd.DataFrame(columns = ['ReceiveDate', 'Subject', 'Sender', 'Importance', "Body"])

    data = dict()

    count = 0
    for messageDetails in message:
        if messageDetails.UnRead==True and messageDetails.SenderEmailAddress.startswith('xxxxxx') and messageDetails.Subject.startswith('xxxxx'):
            df.loc[count,"ReceiveDate"] = (messageDetails.ReceivedTime).strftime("%m/%d/%Y, %H:%M:%S")
            df.loc[count,"Subject"] = (messageDetails.Subject)
            df.loc[count,"Sender"] = (messageDetails.SenderEmailAddress)
            df.loc[count,"Importance"] = (messageDetails.Importance)
            df.loc[count,"Body"] = (messageDetails.Body)
            messageDetails.UnRead =False
            #messageDetails.Move(validated_mails)
            count += 1

    print('loading.....')
    print(df.head(5))

    book = load_workbook("C:/tetx1.xlsx")
    writer = pd.ExcelWriter('C:/tetx1.xlsx', engine='openpyxl', mode="a")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
    print("saving to Excel.....")
    writer.save()
    print("saved..... ")
    print("Thank you very much :)")

 
if __name__ == "__main__":
    run=hello_world()
    run=outlook_email()
